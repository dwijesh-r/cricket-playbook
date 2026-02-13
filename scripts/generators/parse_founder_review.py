#!/usr/bin/env python3
"""
Cricket Playbook - Founder Review Parser
Author: Brock Purdy (Data Pipeline Owner)
Sprint 4.0 - Founder Data Integration

Parses the Founder's IPL 2026 Team Review Excel file into:
1. Canonical JSON (outputs/founder_review/founder_squads_2026.json)
2. Parquet for SQL Lab (scripts/the_lab/dashboard/data/sql_lab/tables/founder_squads_2026.parquet)
3. Updates ipl_2026_squads.csv with authoritative Founder data

First 12 players per team = Founder's Predicted XII.
"""

import csv
import json
import re
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional

from scripts.utils.logging_config import setup_logger

logger = setup_logger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent.parent
DATA_DIR = PROJECT_DIR / "data"
OUTPUT_DIR = PROJECT_DIR / "outputs"
FOUNDER_OUTPUT_DIR = OUTPUT_DIR / "founder_review"
SQL_LAB_DIR = PROJECT_DIR / "scripts" / "the_lab" / "dashboard" / "data" / "sql_lab" / "tables"

REFERENCE_DATE = date.today()  # Age as of commit date

SHEET_TO_ABBREV = {
    "MI": "MI",
    "CSK": "CSK",
    "DC": "DC",
    "GT": "GT",
    "KKR": "KKR",
    "LSG": "LSG",
    "PBKS": "PBKS",
    "RR": "RR",
    "RCB": "RCB",
    "SRH": "SRH",
}

ABBREV_TO_FULL = {
    "MI": "Mumbai Indians",
    "CSK": "Chennai Super Kings",
    "DC": "Delhi Capitals",
    "GT": "Gujarat Titans",
    "KKR": "Kolkata Knight Riders",
    "LSG": "Lucknow Super Giants",
    "PBKS": "Punjab Kings",
    "RR": "Rajasthan Royals",
    "RCB": "Royal Challengers Bengaluru",
    "SRH": "Sunrisers Hyderabad",
}

# Bowling type normalization from Founder's format to our short form
BOWLING_TYPE_MAP = {
    "left-arm orthodox": "LA Orthodox",
    "left-arm seam": "LA Seam",
    "left-arm fast": "LA Fast",
    "left-arm wrist spin": "LA Wrist-spin",
    "left-arm wrist-spin": "LA Wrist-spin",
    "leg-spin": "Leg-spin",
    "off-spin": "Off-spin",
    "wrist-spin": "Wrist-spin",
    "medium-fast": "Fast",
    "fast": "Fast",
    "medium": "Medium",
    "right-arm fast": "Fast",
    "right-arm medium": "Medium",
    "right-arm off-spin": "Off-spin",
    "right-arm leg-spin": "Leg-spin",
}

# Nationality normalization
NATIONALITY_MAP = {
    "india": "IND",
    "australia": "AUS",
    "england": "ENG",
    "south africa": "RSA",
    "new zealand": "NZ",
    "west indies": "WI",
    "sri lanka": "SL",
    "afghanistan": "AFG",
    "bangladesh": "BAN",
    "zimbabwe": "ZIM",
    "pakistan": "PAK",
    "ireland": "IRE",
    "netherlands": "NED",
    "scotland": "SCO",
}

# Batting hand normalization
BATTING_HAND_MAP = {
    "lhb": "Left-hand",
    "rhb": "Right-hand",
    "left-hand": "Left-hand",
    "right-hand": "Right-hand",
}

# Bowling hand normalization
BOWLING_HAND_MAP = {
    "la": "Left-arm",
    "ra": "Right-arm",
    "left-arm": "Left-arm",
    "right-arm": "Right-arm",
}

# Role normalization
ROLE_MAP = {
    "batter": "Batter",
    "bowler": "Bowler",
    "allrounder": "All-rounder",
    "all-rounder": "All-rounder",
    "batting allrounder": "All-rounder",
    "bowling allrounder": "All-rounder",
    "wicketkeeper": "Wicketkeeper",
    "wicketkeeper-batter": "Wicketkeeper",
    "wk": "Wicketkeeper",
}

# Manual name overrides for known mismatches between Excel and CSV
NAME_OVERRIDES = {
    "Vaibhav Sooryavanshi": "Vaibhav Suryavanshi",
    "AM Ghazanfar": "AM Ghazanfar",
    "Vyshak Vijaykumar": "Vijaykumar Vyshak",
    "Tejasvi Dahiya": "Tejasvi Singh",
}


# =============================================================================
# PARSING FUNCTIONS
# =============================================================================


def parse_birthday(raw: Any) -> Optional[date]:
    """Parse birthday from various formats in the Excel file."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        raw = raw.strip()
        # Handle "Sept 8, 1999" format
        for fmt in [
            "%B %d, %Y",
            "%b %d, %Y",
            "%d %B %Y",
            "%d %b %Y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
        ]:
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        # Handle abbreviated month names like "Sept" which strptime doesn't know
        raw_fixed = re.sub(r"\bSept\b", "Sep", raw)
        for fmt in ["%b %d, %Y", "%d %b %Y"]:
            try:
                return datetime.strptime(raw_fixed, fmt).date()
            except ValueError:
                continue
        logger.warning("Could not parse birthday: %s", raw)
    return None


def calculate_age(birthday: Optional[date]) -> Optional[int]:
    """Calculate age in years as of REFERENCE_DATE."""
    if birthday is None:
        return None
    age = REFERENCE_DATE.year - birthday.year
    if (REFERENCE_DATE.month, REFERENCE_DATE.day) < (birthday.month, birthday.day):
        age -= 1
    return age


def normalize_bowling_type(raw: Optional[str]) -> Optional[str]:
    """Normalize bowling type to short form."""
    if not raw:
        return None
    key = raw.strip().lower()
    return BOWLING_TYPE_MAP.get(key, raw.strip())


def normalize_nationality(raw: Optional[str]) -> str:
    """Normalize nationality to 3-letter code."""
    if not raw:
        return "IND"  # Default
    key = raw.strip().lower()
    return NATIONALITY_MAP.get(key, raw.strip()[:3].upper())


def normalize_batting_hand(raw: Optional[str]) -> str:
    """Normalize batting hand."""
    if not raw:
        return "Right-hand"
    key = raw.strip().lower()
    return BATTING_HAND_MAP.get(key, raw.strip())


def normalize_bowling_hand(raw: Optional[str]) -> Optional[str]:
    """Normalize bowling hand/arm."""
    if not raw:
        return None
    key = raw.strip().lower()
    return BOWLING_HAND_MAP.get(key, raw.strip())


def normalize_role(raw: Optional[str]) -> str:
    """Normalize role to our standard."""
    if not raw:
        return "Batter"
    key = raw.strip().lower()
    return ROLE_MAP.get(key, raw.strip())


def load_squad_csv() -> Dict[str, Dict[str, Dict[str, str]]]:
    """Load existing squad CSV indexed by team+player_name."""
    squad_file = DATA_DIR / "ipl_2026_squads.csv"
    squads = {}
    with open(squad_file, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = f"{row['team_name']}|{row['player_name']}"
            squads[key] = dict(row)
    return squads


def fuzzy_match_player(
    founder_name: str,
    team_full: str,
    csv_players: Dict[str, Dict[str, str]],
) -> Optional[str]:
    """Match Founder player name to CSV player using exact then fuzzy matching."""
    # Check manual overrides first
    search_name = NAME_OVERRIDES.get(founder_name, founder_name)

    # Exact match
    exact_key = f"{team_full}|{search_name}"
    if exact_key in csv_players:
        return exact_key

    # Fuzzy match within team
    best_match = None
    best_ratio = 0.0
    for key, row in csv_players.items():
        if row["team_name"] != team_full:
            continue
        ratio = SequenceMatcher(None, search_name.lower(), row["player_name"].lower()).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = key

    if best_ratio >= 0.75:
        return best_match

    return None


def load_contracts() -> Dict[str, float]:
    """Load player contracts for price data."""
    contracts = {}
    contract_file = DATA_DIR / "ipl_2026_player_contracts.csv"
    if contract_file.exists():
        with open(contract_file, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = f"{row['team_name']}|{row['player_name']}"
                contracts[key] = float(row["price_cr"])
    return contracts


def load_experience() -> Dict[str, Dict[str, Any]]:
    """Load IPL experience data."""
    experience = {}
    exp_file = OUTPUT_DIR / "team" / "ipl_2026_squad_experience.csv"
    if exp_file.exists():
        with open(exp_file, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = f"{row['team_name']}|{row['player_name']}"
                bat_innings = (
                    float(row["ipl_batting_innings"]) if row.get("ipl_batting_innings") else 0
                )
                bowl_matches = (
                    float(row["ipl_bowling_matches"]) if row.get("ipl_bowling_matches") else 0
                )
                experience[key] = {
                    "ipl_matches": int(max(bat_innings, bowl_matches)),
                    "ipl_batting_runs": int(float(row.get("ipl_batting_runs", 0) or 0)),
                    "ipl_batting_sr": float(row.get("ipl_batting_sr", 0) or 0),
                    "ipl_bowling_wickets": int(float(row.get("ipl_bowling_wickets", 0) or 0)),
                    "ipl_bowling_economy": float(row.get("ipl_bowling_economy", 0) or 0),
                }
    return experience


def parse_excel() -> List[Dict[str, Any]]:
    """Parse all sheets from the Founder Review Excel file."""
    import openpyxl

    excel_path = PROJECT_DIR / "reviews" / "founder" / "IPL 2026 Teams - Founder Review.xlsx"
    logger.info("Parsing Founder Review: %s", excel_path)

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    csv_players = load_squad_csv()
    contracts = load_contracts()
    experience = load_experience()

    all_players = []
    unmatched = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_ABBREV:
            logger.warning("Skipping unknown sheet: %s", sheet_name)
            continue

        team_abbrev = SHEET_TO_ABBREV[sheet_name]
        team_full = ABBREV_TO_FULL[team_abbrev]
        ws = wb[sheet_name]

        logger.info("Processing %s (%s)...", team_full, team_abbrev)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Skip header row
        player_count = 0
        for row in rows[1:]:
            # Skip empty rows
            if not row[0] or not row[1]:
                continue

            squad_number = int(row[0])
            player_name = str(row[1]).strip()
            batting_hand = normalize_batting_hand(row[2])
            bowling_hand = normalize_bowling_hand(row[3])
            bowling_type = normalize_bowling_type(row[4])
            role = normalize_role(row[5])
            birthday = parse_birthday(row[6])
            age = calculate_age(birthday)
            nationality = normalize_nationality(row[7])
            notes = str(row[8]).strip() if row[8] else None

            # Match to CSV
            csv_key = fuzzy_match_player(player_name, team_full, csv_players)
            player_id = None
            if csv_key:
                player_id = csv_players[csv_key]["player_id"]
            else:
                unmatched.append(f"{team_abbrev}: {player_name}")
                logger.warning("UNMATCHED: %s - %s", team_abbrev, player_name)

            # Get price and experience
            lookup_key = csv_key or f"{team_full}|{player_name}"
            price_cr = contracts.get(lookup_key, 0.0)
            exp_data = experience.get(lookup_key, {})

            is_predicted_xii = squad_number <= 12

            player_data = {
                "team_abbrev": team_abbrev,
                "team_name": team_full,
                "squad_number": squad_number,
                "player_name": player_name,
                "player_id": player_id,
                "batting_hand": batting_hand,
                "bowling_arm": bowling_hand,
                "bowling_type": bowling_type,
                "role": role,
                "birthday": birthday.isoformat() if birthday else None,
                "age": age,
                "nationality": nationality,
                "notes": notes,
                "is_predicted_xii": is_predicted_xii,
                "batting_position": squad_number if is_predicted_xii else None,
                "price_cr": price_cr,
                "ipl_matches": exp_data.get("ipl_matches", 0),
                "ipl_batting_runs": exp_data.get("ipl_batting_runs", 0),
                "ipl_batting_sr": exp_data.get("ipl_batting_sr", 0.0),
                "ipl_bowling_wickets": exp_data.get("ipl_bowling_wickets", 0),
                "ipl_bowling_economy": exp_data.get("ipl_bowling_economy", 0.0),
            }

            all_players.append(player_data)
            player_count += 1

            if player_count >= 25:
                break

        logger.info(
            "  %s: %d players parsed, XII: %s",
            team_abbrev,
            player_count,
            ", ".join(
                p["player_name"] for p in all_players[-player_count:] if p["is_predicted_xii"]
            )[:80]
            + "...",
        )

    logger.info("Total: %d players parsed, %d unmatched", len(all_players), len(unmatched))
    if unmatched:
        logger.warning("Unmatched players: %s", unmatched)

    wb.close()
    return all_players


# =============================================================================
# OUTPUT FUNCTIONS
# =============================================================================


def save_json(players: List[Dict[str, Any]]) -> Path:
    """Save canonical JSON."""
    FOUNDER_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = {
        "generated_at": datetime.now().isoformat(),
        "reference_date": REFERENCE_DATE.isoformat(),
        "source": "reviews/founder/IPL 2026 Teams - Founder Review.xlsx",
        "total_players": len(players),
        "teams": {},
    }

    for p in players:
        team = p["team_abbrev"]
        if team not in output["teams"]:
            output["teams"][team] = {
                "team_name": p["team_name"],
                "team_abbrev": team,
                "players": [],
                "predicted_xii": [],
            }
        output["teams"][team]["players"].append(p)
        if p["is_predicted_xii"]:
            output["teams"][team]["predicted_xii"].append(p["player_name"])

    out_path = FOUNDER_OUTPUT_DIR / "founder_squads_2026.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    logger.info("Saved JSON: %s (%d players)", out_path, len(players))
    return out_path


def save_parquet(players: List[Dict[str, Any]]) -> Path:
    """Save Parquet for SQL Lab."""
    import pandas as pd

    SQL_LAB_DIR.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(players)
    # Select and order columns for SQL Lab
    columns = [
        "team_abbrev",
        "team_name",
        "squad_number",
        "player_name",
        "player_id",
        "batting_hand",
        "bowling_arm",
        "bowling_type",
        "role",
        "age",
        "nationality",
        "notes",
        "is_predicted_xii",
        "batting_position",
        "price_cr",
        "ipl_matches",
        "ipl_batting_runs",
        "ipl_batting_sr",
        "ipl_bowling_wickets",
        "ipl_bowling_economy",
    ]
    df = df[[c for c in columns if c in df.columns]]

    out_path = SQL_LAB_DIR / "founder_squads_2026.parquet"
    df.to_parquet(out_path, index=False)
    logger.info("Saved Parquet: %s (%d rows)", out_path, len(df))
    return out_path


def update_squad_csv(players: List[Dict[str, Any]]) -> int:
    """Update ipl_2026_squads.csv with Founder data (nationality, bowling_type, age, role, batting_hand, bowling_arm)."""
    squad_file = DATA_DIR / "ipl_2026_squads.csv"

    # Read existing CSV
    with open(squad_file, "r") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    # Build lookup from Founder data by player_id
    founder_lookup = {}
    for p in players:
        if p["player_id"]:
            founder_lookup[p["player_id"]] = p

    # Update rows
    updated_count = 0
    for row in rows:
        pid = row["player_id"]
        if pid in founder_lookup:
            fp = founder_lookup[pid]
            # Update ONLY these columns from Founder
            row["nationality"] = fp["nationality"]
            row["bowling_type"] = fp["bowling_type"] or row.get("bowling_type", "")
            row["bowling_arm"] = fp["bowling_arm"] or row.get("bowling_arm", "")
            row["age"] = str(fp["age"]) if fp["age"] is not None else row.get("age", "")
            row["role"] = fp["role"]
            row["batting_hand"] = fp["batting_hand"]
            updated_count += 1

    # Write back
    with open(squad_file, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    logger.info("Updated %d players in %s", updated_count, squad_file)
    return updated_count


# =============================================================================
# IPL_2026_Final_Version VIEW
# =============================================================================


def create_final_version_view(players: List[Dict[str, Any]]) -> str:
    """
    Generate SQL for IPL_2026_Final_Version view that joins
    founder_squads_2026 with ipl_2026_squads and ipl_2026_contracts.
    """
    view_sql = """-- IPL_2026_Final_Version: Authoritative merged view of Founder Review + Squad + Contracts + Experience
CREATE OR REPLACE VIEW IPL_2026_Final_Version AS
SELECT
    fs.team_abbrev,
    fs.team_name,
    fs.squad_number,
    fs.player_name,
    fs.player_id,
    fs.role,
    fs.batting_hand,
    fs.bowling_arm,
    fs.bowling_type,
    fs.age,
    fs.nationality,
    fs.is_predicted_xii,
    fs.batting_position,
    fs.notes AS founder_notes,
    fs.price_cr,
    fs.ipl_matches,
    fs.ipl_batting_runs,
    fs.ipl_batting_sr,
    fs.ipl_bowling_wickets,
    fs.ipl_bowling_economy,
    sq.batter_classification,
    sq.bowler_classification,
    sq.batter_tags,
    sq.bowler_tags,
    sq.is_captain
FROM founder_squads_2026 fs
LEFT JOIN ipl_2026_squads sq ON fs.player_id = sq.player_id;
"""
    return view_sql


# =============================================================================
# MAIN
# =============================================================================


def main():
    logger.info("=" * 60)
    logger.info("FOUNDER REVIEW PARSER - IPL 2026")
    logger.info("Reference date for age: %s", REFERENCE_DATE)
    logger.info("=" * 60)

    # Parse Excel
    players = parse_excel()

    if not players:
        logger.error("No players parsed!")
        return 1

    # Save outputs
    save_json(players)
    save_parquet(players)

    # Update squad CSV
    updated = update_squad_csv(players)
    logger.info("Updated %d players in squad CSV", updated)

    # Generate Final Version view SQL
    view_sql = create_final_version_view(players)
    view_file = FOUNDER_OUTPUT_DIR / "ipl_2026_final_version_view.sql"
    with open(view_file, "w") as f:
        f.write(view_sql)
    logger.info("Saved view SQL: %s", view_file)

    # Summary
    logger.info("=" * 60)
    logger.info("SUMMARY")
    logger.info("Total players: %d", len(players))
    matched = sum(1 for p in players if p["player_id"])
    logger.info("Matched to CSV: %d/%d", matched, len(players))
    xii_count = sum(1 for p in players if p["is_predicted_xii"])
    logger.info("Predicted XII selections: %d (should be 120)", xii_count)

    for team in sorted(ABBREV_TO_FULL.keys()):
        team_players = [p for p in players if p["team_abbrev"] == team]
        team_xii = [p for p in team_players if p["is_predicted_xii"]]
        team_unmatched = [p for p in team_players if not p["player_id"]]
        logger.info(
            "  %s: %d players, XII=%d, unmatched=%d",
            team,
            len(team_players),
            len(team_xii),
            len(team_unmatched),
        )

    logger.info("=" * 60)
    return 0


if __name__ == "__main__":
    main()
